class addsimulation:
    
    def set_OutputEfficientFrontierObject(self,OutputEfficientFrontierObject):
        self._OutputEfficientFrontierObject = OutputEfficientFrontierObject
    def get_OutputEfficientFrontierObject(self):
        return self._OutputEfficientFrontierObject
    OutputEfficientFrontierObject = property(get_OutputEfficientFrontierObject, set_OutputEfficientFrontierObject)
    
    def __init__(self,symbols,startdate,enddate,permutations,annualized_or_cumulative):
        print 'initialized addsimulationtoworkbook'
        import compileworkbook as cwbk
        o = cwbk.compileclass(symbols,startdate,enddate,permutations,annualized_or_cumulative)
        path_to_workbook = o.PathnameToCompiledWorkbook
        #path_to_workbook = 'C:\\Batches\\GitStuff\\modern-portfolio-theory\\cache\\20170921102057 compiled.xlsx'
        from os.path import basename
        import os
        rootpath = os.path.dirname(os.path.abspath(path_to_workbook))
        filename = basename(path_to_workbook)
        fileroot = filename.split(' ')[0]
        newfilename = fileroot + ' simulated.xlsx'
        new_path_to_workbook = os.path.join(rootpath, newfilename)
        from openpyxl import load_workbook
        from openpyxl.chart import (
            ScatterChart,
            Reference,
            Series,
        )
        
        wb = load_workbook(filename = path_to_workbook)
        current_wks_name = 'permutations'
        print '-----',current_wks_name,'------------'
        wks = wb[current_wks_name]
        col_portfoliostandarddeviation = 0
        col_portfolioreturn = 0
        row_max = 0
        for row_cells in wks.iter_rows(min_row=1, max_row=1):
            for cell in row_cells:
                print cell.col_idx, cell.column, cell.row, cell.value
                if cell.value == 'portfoliostandarddeviation':
                    col_portfoliostandarddeviation = cell.col_idx
                if cell.value == 'portfolioreturn':
                    col_portfolioreturn = cell.col_idx
        for col_cells in wks.iter_rows(min_col=1, max_col=1):
            for cell in col_cells:
                #print cell.col_idx, cell.column, cell.row, cell.value
                row_max = cell.row
        
        chart = ScatterChart()
        chart.title = "Scatter Chart"
        chart.style = 5
        chart.charttype = -4169 

        chart.x_axis.title = 'Risk'
        chart.y_axis.title = 'Return'

        xvalues = Reference(wks, min_col=col_portfoliostandarddeviation, min_row=2, max_row=row_max)
        yvalues = Reference(wks, min_col=col_portfolioreturn, min_row=1, max_row=row_max)
        series = Series(yvalues, xvalues, title_from_data=True)
        chart.series.append(series)
        wks.add_chart(chart, "N10")
        
        
        current_wks_name = 'aggregatedpricechangereturns'
        print '-----',current_wks_name,'------------' 
        wks = wb[current_wks_name]
        col_portfoliostandarddeviation = 0
        col_portfolioreturn = 0
        row_max = 0
        col_max = 0
        for row_cells in wks.iter_rows(min_row=1, max_row=1):
            for cell in row_cells:
                print cell.col_idx, cell.column, cell.row, cell.value
                if cell.value == 'symbol':
                    col_symbol = cell.col_idx
                if cell.value == 'cumulative_return':
                    col_cumulative_return = cell.col_idx
                if cell.value == 'random_return':
                    col_random_return = cell.col_idx
                col_max = cell.col_idx
        for col_cells in wks.iter_rows(min_col=1, max_col=1,min_row=2):
            for cell in col_cells:
                #print cell.col_idx, cell.column, cell.row, cell.value
                row_max = cell.row

        wb.create_sheet('simulation')
        wks_simulation = wb['simulation']

        wb.save(new_path_to_workbook)
##        try:
##            wb.create_sheet('simulation')
##            wks_simulation = wb['simulation']
##            wb.save('C:\\Batches\\GitStuff\\modern-portfolio-theory\\cache\\20170921102057 simulated.xlsx')
##            simulation_row_offset = 7
##            print 'simulation_row_offset',simulation_row_offset
##            for row_cells in wks.iter_rows(min_row=1, max_row=row_max):
##                for c in row_cells:
##                    wks_simulation.cell(row=c.row, column=c.column).value = c.value
##                    print 'got here 1',c,wks_simulation.cell(row=c.row, column=c.column).value
##            for col_cells in wks.iter_rows(min_col=col_symbol, max_col=col_symbol,min_row=1):
##                for c in col_cells:
##                    wks_simulation.cell(row=c.row, column=col_max+2).value = wks.cell(row=c.row, column=col_symbol).value
##                    wks_simulation.cell(row=c.row, column=col_max+3).value = wks.cell(row=c.row, column=col_random_return).value
##                    print 'got here 2',c,wks_simulation.cell(row=c.row, column=col_max+3).value
##                    if c.row == 1:
##                        print 'got here 3',c
##                        wks_simulation.cell(row=c.row, column=col_max+2).value = 'Symbol'
##                        wks_simulation.cell(row=c.row, column=col_max+3).value = 'Forecast Return'
##                        wks_simulation.cell(row=c.row, column=col_max+4).value = 'Top Constraint'
##                        wks_simulation.cell(row=c.row, column=col_max+5).value = 'Bottom Constraint'
##                        wks_simulation.cell(row=c.row, column=col_max+6).value = 'Simulated Weights'
##                        wks_simulation.cell(row=c.row, column=col_max+7).value = 'Equal Weights'
##        except:
##            pass
##        print 'got here 4'
##        wb.save('C:\\Batches\\GitStuff\\modern-portfolio-theory\\cache\\20170921102057 simulatedxx.xlsx')
        

if __name__=='__main__':
##spy
##iwm
##mdy
##'spy',
##'iwm',
##'mdy'



##                'dspg',
##                'bud',
##                'px',
##                'unp',
##                'lbrdk',
##                'chtr',
##                'avgo',
##                'stz',
##                'mar',
##                'mgm',
##                'lsxma',
##                'lyb',
##                'goog',
##                'sbac',
##                'fb',
##                'low',
##                'aap',
##                'pcln',
##                'wlk',
##                'crc',
##                'adsk',
##                'cmcsa',
##                'dltr',
##                'mpc'



        o = addsimulation(
                    symbols = ['MAR', 'MON', 'NOV', 'A', 'AAL', 'AAP', 'AAPL', 'ABBV', 'ABC', 'ABT', 'ACN', 'ADBE', 'ADI', 'ADM', 'ADP', 'ADS', 'ADSK', 'AEE', 'AEP', 'AES', 'AET', 'AFL', 'AGN', 'AIG', 'AIV', 'AIZ', 'AJG', 'AKAM', 'ALB', 'ALGN', 'ALK', 'ALL', 'ALLE', 'ALXN', 'AMAT', 'AMD', 'AME', 'AMG', 'AMGN', 'AMP', 'AMT', 'AMZN', 'ANDV', 'ANSS', 'ANTM', 'AON', 'AOS', 'APA', 'APC', 'APD', 'APH', 'ARE', 'ARNC', 'ATVI', 'AVB', 'AVGO', 'AVY', 'AWK', 'AXP', 'AYI', 'AZO', 'BA', 'BAC', 'BAX', 'BBT', 'BBY', 'BCR', 'BDX', 'BEN', 'BF.B', 'BHF', 'BHGE', 'BIIB', 'BK', 'BLK', 'BLL', 'BMY', 'BRK.B', 'BSX', 'BWA', 'BXP', 'C', 'CA', 'CAG', 'CAH', 'CAT', 'CB', 'CBG', 'CBOE', 'CBS', 'CCI', 'CCL', 'CDNS', 'CELG', 'CERN', 'CF', 'CFG', 'CHD', 'CHK', 'CHRW', 'CHTR', 'CI', 'CINF', 'CL', 'CLX', 'CMA', 'CMCSA', 'CME', 'CMG', 'CMI', 'CMS', 'CNC', 'CNP', 'COF', 'COG', 'COH', 'COL', 'COO', 'COP', 'COST', 'COTY', 'CPB', 'CRM', 'CSCO', 'CSRA', 'CSX', 'CTAS', 'CTL', 'CTSH', 'CTXS', 'CVS', 'CVX', 'CXO', 'D', 'DAL', 'DE', 'DFS', 'DG', 'DGX', 'DHI', 'DHR', 'DIS', 'DISCA', 'DISCK', 'DISH', 'DLPH', 'DLR', 'DLTR', 'DOV', 'DPS', 'DRE', 'DRI', 'DTE', 'DUK', 'DVA', 'DVN', 'DWDP', 'DXC', 'EA', 'EBAY', 'ECL', 'ED', 'EFX', 'EIX', 'EL', 'EMN', 'EMR', 'EOG', 'EQIX', 'EQR', 'EQT', 'ES', 'ESRX', 'ESS', 'ETFC', 'ETN', 'ETR', 'EVHC', 'EW', 'EXC', 'EXPD', 'EXPE', 'EXR', 'F', 'FAST', 'FB', 'FBHS', 'FCX', 'FDX', 'FE', 'FFIV', 'FIS', 'FISV', 'FITB', 'FL', 'FLIR', 'FLR', 'FLS', 'FMC', 'FOX', 'FOXA', 'FRT', 'FTI', 'FTV', 'GD', 'GE', 'GGP', 'GILD', 'GIS', 'GLW', 'GM', 'GOOG', 'GOOGL', 'GPC', 'GPN', 'GPS', 'GRMN', 'GS', 'GT', 'GWW', 'HAL', 'HAS', 'HBAN', 'HBI', 'HCA', 'HCN', 'HCP', 'HD', 'HES', 'HIG', 'HLT', 'HOG', 'HOLX', 'HON', 'HP', 'HPE', 'HPQ', 'HRB', 'HRL', 'HRS', 'HSIC', 'HST', 'HSY', 'HUM', 'IBM', 'ICE', 'IDXX', 'IFF', 'ILMN', 'INCY', 'INFO', 'INTC', 'INTU', 'IP', 'IPG', 'IR', 'IRM', 'ISRG', 'IT', 'ITW', 'IVZ', 'JBHT', 'JCI', 'JEC', 'JNJ', 'JNPR', 'JPM', 'JWN', 'K', 'KEY', 'KHC', 'KIM', 'KLAC', 'KMB', 'KMI', 'KMX', 'KO', 'KORS', 'KR', 'KSS', 'KSU', 'L', 'LB', 'LEG', 'LEN', 'LH', 'LKQ', 'LLL', 'LLY', 'LMT', 'LNC', 'LNT', 'LOW', 'LRCX', 'LUK', 'LUV', 'LVLT', 'LYB', 'M', 'MA', 'MAA', 'MAC', 'MAS', 'MAT', 'MCD', 'MCHP', 'MCK', 'MCO', 'MDLZ', 'MDT', 'MET', 'MGM', 'MHK', 'MKC', 'MLM', 'MMC', 'MMM', 'MNST', 'MO', 'MOS', 'MPC', 'MRK', 'MRO', 'MS', 'MSFT', 'MSI', 'MTB', 'MTD', 'MU', 'MYL', 'NAVI', 'NBL', 'NDAQ', 'NEE', 'NEM', 'NFLX', 'NFX', 'NI', 'NKE', 'NLSN', 'NOC', 'NRG', 'NSC', 'NTAP', 'NTRS', 'NUE', 'NVDA', 'NWL', 'NWS', 'NWSA', 'O', 'OKE', 'OMC', 'ORCL', 'ORLY', 'OXY', 'PAYX', 'PBCT', 'PCAR', 'PCG', 'PCLN', 'PDCO', 'PEG', 'PEP', 'PFE', 'PFG', 'PG', 'PGR', 'PH', 'PHM', 'PKG', 'PKI', 'PLD', 'PM', 'PNC', 'PNR', 'PNW', 'PPG', 'PPL', 'PRGO', 'PRU', 'PSA', 'PSX', 'PVH', 'PWR', 'PX', 'PXD', 'PYPL', 'Q', 'QCOM', 'QRVO', 'RCL', 'RE', 'REG', 'REGN', 'RF', 'RHI', 'RHT', 'RJF', 'RL', 'RMD', 'ROK', 'ROP', 'ROST', 'RRC', 'RSG', 'RTN', 'SBAC', 'SBUX', 'SCG', 'SCHW', 'SEE', 'SHW', 'SIG', 'SJM', 'SLB', 'SLG', 'SNA', 'SNI', 'SNPS', 'SO', 'SPG', 'SPGI', 'SPLS', 'SRCL', 'SRE', 'STI', 'STT', 'STX', 'STZ', 'SWK', 'SWKS', 'SYF', 'SYK', 'SYMC', 'SYY', 'T', 'TAP', 'TDG', 'TEL', 'TGT', 'TIF', 'TJX', 'TMK', 'TMO', 'TRIP', 'TROW', 'TRV', 'TSCO', 'TSN', 'TSS', 'TWX', 'TXN', 'TXT', 'UA', 'UAA', 'UAL', 'UDR', 'UHS', 'ULTA', 'UNH', 'UNM', 'UNP', 'UPS', 'URI', 'USB', 'UTX', 'V', 'VAR', 'VFC', 'VIAB', 'VLO', 'VMC', 'VNO', 'VRSK', 'VRSN', 'VRTX', 'VTR', 'VZ', 'WAT', 'WBA', 'WDC', 'WEC', 'WFC', 'WHR', 'WLTW', 'WM', 'WMB', 'WMT', 'WRK', 'WU', 'WY', 'WYN', 'WYNN', 'XEC', 'XEL', 'XL', 'XLNX', 'XOM', 'XRAY', 'XRX', 'XYL', 'YUM', 'ZBH', 'ZION', 'ZTS']
                    #symbols = ['LAZ', 'LMT', 'RTN', 'MAS', 'AMAT', 'INTC', 'LPX', 'GRMN', 'PCLN', 'KSS', 'JWN', 'M', 'GPS', 'LOW', 'PEP', 'CVS', 'CL', 'KMB', 'MO', 'PM', 'CVX', 'BAC', 'BEN', 'MS', 'AXP', 'CELG', 'AMGN', 'JNJ', 'LLY', 'MMM', 'UNP', 'CSCO', 'SWKS', 'CA', 'STX', 'LYB', 'APD', 'T', 'TGT', 'HD', 'ETR', 'AES', 'HOG', 'F', 'GPC', 'LEG', 'WHR', 'NWL', 'TRIP', 'HAS', 'BC', 'CMCSA', 'DIS', 'VIA', 'DISH', 'NWS', 'PAG', 'CRI', 'COLM', 'SKX', 'NKE', 'TAP', 'CASY', 'HRL', 'HAIN', 'SJM', 'ADM', 'KHC', 'MDLZ', 'FTI', 'SLB', 'NFX', 'KMI', 'CXO', 'MUR', 'WPX', 'EGN', 'XOM', 'LNG', 'FCNCA', 'LUK', 'Y', 'WTM', 'AXS', 'ALKS', 'MDT', 'XRAY', 'CAH', 'MD', 'PDCO', 'UHS', 'AGN', 'ARNC', 'UAL', 'AAL', 'GE', 'SNA', 'WAB', 'FLS', 'VRSK', 'GWR', 'GWW', 'VSAT', 'AVT', 'TWTR', 'AMD', 'QCOM', 'FSLR', 'OTEX', 'NUAN', 'HPE', 'RPM', 'MLM', 'VMC', 'SEE', 'SON', 'HHC', 'LVLT', 'S', 'JLL']
                    ,  startdate = '2017-03-01'
                    ,  enddate = '2017-09-30'
                    ,  permutations = 1000
                    ,  annualized_or_cumulative = 'cumulative'
                  )
