import pandas as pd
class compileclass:
    
    def set_OutputEfficientFrontierObject(self,OutputEfficientFrontierObject):
        self._OutputEfficientFrontierObject = OutputEfficientFrontierObject
    def get_OutputEfficientFrontierObject(self):
        return self._OutputEfficientFrontierObject
    OutputEfficientFrontierObject = property(get_OutputEfficientFrontierObject, set_OutputEfficientFrontierObject)

    def set_PathnameToCompiledWorkbook(self,PathnameToCompiledWorkbook):
        self._PathnameToCompiledWorkbook = PathnameToCompiledWorkbook
    def get_PathnameToCompiledWorkbook(self):
        return self._PathnameToCompiledWorkbook
    PathnameToCompiledWorkbook = property(get_PathnameToCompiledWorkbook, set_PathnameToCompiledWorkbook)

    
    def __init__(self,symbols_and_signs_dataframe,startdate,enddate,permutations,annualized_or_cumulative):
        import config
        import mytools
        import datetime
        import os
        import outputefficientfrontier as oef
        a = list([symbols_and_signs_dataframe.index.tolist(),symbols_and_signs_dataframe['longshort'].tolist()])
        symbols_and_signs_list = zip(*a)
        
        #print symbols_and_signs_list
        #stop
        #symbols = [i[0] for i in symbols_and_signs_list]
        #longshortlist = [i[1] for i in symbols_and_signs_list]
        #symbols_and_signs_list = xxx
        o = oef.output(symbols_and_signs_list,startdate,enddate,permutations,annualized_or_cumulative)
        d = o.DictionaryOfOutputFiles
        for k,v in d.items():
            #print k, v
            (f_path, filename) = os.path.split(v)
            date14 = filename.split(' ')[0]
        print date14
        mycachefolder = config.mycachefolder
        mytools.general().make_sure_path_exists(mycachefolder)    
        #date14 = str(datetime.datetime.now().strftime("%Y%m%d%H%M%S"))
        cachedfilepathname = mycachefolder
        cachedfilepathname = os.path.join(cachedfilepathname,date14 + ' compiled.xlsx')

        import xlsxwriter
        import glob
        import csv
        import numbers
        workbook = xlsxwriter.Workbook(cachedfilepathname, {'strings_to_numbers': True}) 
        #for filename in glob.glob("*.csv"):
        for k,v in d.items():
            #print k, v
            (f_path, filename) = os.path.split(v)
            worksheetname = str(filename.split('.')[0])
            worksheetname = worksheetname.split(' ')[1]
            ws = workbook.add_worksheet(worksheetname)
            spamReader = csv.reader(open(v, 'rb'), delimiter=',')
            row_count = 0
            #print filename
            for row in spamReader:
                for col in range(len(row)):
                    n = row[col]
                    ws.write(row_count,col,n)
                row_count +=1

        workbook.close()
        self.PathnameToCompiledWorkbook = cachedfilepathname
        print 'you can find your compileclass file here: ',cachedfilepathname

    def improveworkbook(self,symbols_and_signs_dataframe,compiledfilepath):

        self.PathnameToCompiledWorkbook = compiledfilepath
        
        import config
        import mytools
        import openpyxl as pyxl
        import itertools
        wb = pyxl.load_workbook(compiledfilepath)
        df = symbols_and_signs_dataframe
        symbols = list(df.index)
        #print symbols
        #stop        
        for sh in wb.sheetnames:
            print sh
        ws = wb["aggregatedpricechangereturns"]
        ws = wb.copy_worksheet(ws)
        ws.title = 'simulation'
        
        for row in ws.iter_rows('B{}:B{}'.format(ws.min_row,ws.max_row)):
            for cell in row:
                if cell.row == 1:
                    ws['N'+str(cell.row)].value = "Long/Short"
                    ws['U'+str(cell.row)].value = "Forecast Return"
                    ws['V'+str(cell.row)].value = "Given Weights"
                    ws['W'+str(cell.row)].value = "Equal Weights"
                    ws['X'+str(cell.row)].value = "Optimal Weights"
                    ws['Y'+str(cell.row)].value = "Top Constraint"
                    ws['Z'+str(cell.row)].value = "Bottom Constraint"
                    ws['AA'+str(cell.row)].value = "Market Value (Optimal)"
                    ws['AB'+str(cell.row)].value = "Shares"
                else:
                    print cell.row, cell.column,  cell.value, df.loc[cell.value]['longshort']
                    ws['N'+str(cell.row)].value = df.loc[cell.value]['longshort']
                    ws['U'+str(cell.row)].value = df.loc[cell.value]['forecastreturn']
                    ws['V'+str(cell.row)].value = df.loc[cell.value]['givenweight']
                    ws['Y'+str(cell.row)].value = df.loc[cell.value]['topconstraint']
                    ws['Z'+str(cell.row)].value = df.loc[cell.value]['bottomconstraint']
                    ws['W'+str(cell.row)].value = 1.0/float(len(df))
                    
        wb.save(compiledfilepath)
        wb.close()
        print 'you can find your compileclass file here (PathnameToCompiledWorkbook): ',compiledfilepath

if __name__=='__main__':
    lst0 = []
    lst0.append({'ticker':'MRK','longshort':'L','givenweight':0.2, 'forecastreturn':0.05,'topconstraint':0.05,'bottomconstraint':0.01})
    lst0.append({'ticker':'THO','longshort':'L','givenweight':0.2, 'forecastreturn':0.05,'topconstraint':0.05,'bottomconstraint':0.01})
    lst0.append({'ticker':'ALGN','longshort':'S','givenweight':0.2, 'forecastreturn':-0.05,'topconstraint':-0.005,'bottomconstraint':-0.02})
    lst0.append({'ticker':'CELG','longshort':'S','givenweight':0.2, 'forecastreturn':-0.05,'topconstraint':-0.005,'bottomconstraint':-0.02})
    lst0.append({'ticker':'MSFT','longshort':'L','givenweight':0.1, 'forecastreturn':0.05,'topconstraint':0.05,'bottomconstraint':0.01})
    lst0.append({'ticker':'FB','longshort':'L','givenweight':0.1, 'forecastreturn':0.05,'topconstraint':0.05,'bottomconstraint':0.01})
    df_symbols_and_signs = pd.DataFrame(lst0)
    df_symbols_and_signs = df_symbols_and_signs.set_index('ticker',drop=True)
    print df_symbols_and_signs
    #stop

    o = compileclass(
                symbols_and_signs_dataframe = df_symbols_and_signs
                ,  startdate = '2017-02-28'
                ,  enddate = '2017-09-30'
                ,  permutations = 500
                ,  annualized_or_cumulative = 'cumulative'
              )
    
    o.improveworkbook(symbols_and_signs_dataframe = df_symbols_and_signs,compiledfilepath=o.PathnameToCompiledWorkbook)
