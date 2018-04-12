import pandas as pd
class compileclass:
    
    def set_OutputEfficientFrontierObject(self,OutputEfficientFrontierObject):
        self._OutputEfficientFrontierObject = OutputEfficientFrontierObject
    def get_OutputEfficientFrontierObject(self):
        return self._OutputEfficientFrontierObject
    OutputEfficientFrontierObject = property(get_OutputEfficientFrontierObject, set_OutputEfficientFrontierObject)

    def set_PathnameToCompiledWorkbook(self,PathnameToCompiledWorkbook):
        self._PathnameToCompiledWorkbook = PathnameToCompiledWorkbook
    def get_PathnameToCompiledWorkbook(self):
        return self._PathnameToCompiledWorkbook
    PathnameToCompiledWorkbook = property(get_PathnameToCompiledWorkbook, set_PathnameToCompiledWorkbook)

    def __init__(self,symbols_and_signs_dataframe,startdate,enddate,permutations,annualized_or_cumulative):
        self.execute(symbols_and_signs_dataframe,startdate,enddate,permutations,annualized_or_cumulative)
        print('compileworkbook initialized')
              
                    
        
    def execute(self,symbols_and_signs_dataframe,startdate,enddate,permutations,annualized_or_cumulative):
        import config
        import mytools
        import datetime
        import os
        import outputefficientfrontier as oef
        a = list([symbols_and_signs_dataframe.index.tolist(),symbols_and_signs_dataframe['longshort'].tolist()])
        symbols_and_signs_list = zip(*a)
        
        #print symbols_and_signs_list
        #stop
        #symbols = [i[0] for i in symbols_and_signs_list]
        #longshortlist = [i[1] for i in symbols_and_signs_list]
        #symbols_and_signs_list = xxx
        o = oef.output(symbols_and_signs_list,startdate,enddate,permutations,annualized_or_cumulative)
        d = o.DictionaryOfOutputFiles
        for k,v in d.items():
            #print k, v
            (f_path, filename) = os.path.split(v)
            date14 = filename.split(' ')[0]
        print date14
        mycachefolder = config.mycachefolder
        mytools.general().make_sure_path_exists(mycachefolder)    
        #date14 = str(datetime.datetime.now().strftime("%Y%m%d%H%M%S"))
        cachedfilepathname = mycachefolder
        cachedfilepathname = os.path.join(cachedfilepathname,date14 + ' compiled.xlsx')

        import xlsxwriter
        import glob
        import csv
        import numbers
        workbook = xlsxwriter.Workbook(cachedfilepathname, {'strings_to_numbers': True}) 
        #for filename in glob.glob("*.csv"):
        for k,v in d.items():
            #print k, v
            (f_path, filename) = os.path.split(v)
            worksheetname = str(filename.split('.')[0])
            worksheetname = worksheetname.split(' ')[1]
            ws = workbook.add_worksheet(worksheetname)
            spamReader = csv.reader(open(v, 'rb'), delimiter=',')
            row_count = 0
            #print filename
            for row in spamReader:
                for col in range(len(row)):
                    n = row[col]
                    ws.write(row_count,col,n)
                row_count +=1

        workbook.close()
        self.PathnameToCompiledWorkbook = cachedfilepathname
        print 'you can find your compileclass file here: ',cachedfilepathname

    def improveworkbook(self,symbols_and_signs_dataframe,compiledfilepath):

        self.PathnameToCompiledWorkbook = compiledfilepath
        
        import config
        import mytools
        import openpyxl as pyxl
        import itertools
        wb = pyxl.load_workbook(compiledfilepath)
        df = symbols_and_signs_dataframe
        symbols = list(df.index)
        #print symbols
        #stop        
        for sh in wb.sheetnames:
            print sh
        ws = wb["aggregatedpricechangereturns"]
        ws = wb.copy_worksheet(ws)
        ws.title = 'simulation'
        
        for row in ws.iter_rows('B{}:B{}'.format(ws.min_row,ws.max_row)):
            for cell in row:
                if cell.row == 1:
                    ws['N'+str(cell.row)].value = "Long/Short"
                    ws['U'+str(cell.row)].value = "Forecast Return"
                    ws['V'+str(cell.row)].value = "Given Weights"
                    ws['W'+str(cell.row)].value = "Equal Weights"
                    ws['X'+str(cell.row)].value = "Optimal Weights"
                    ws['Y'+str(cell.row)].value = "Top Constraint"
                    ws['Z'+str(cell.row)].value = "Bottom Constraint"
                    ws['AA'+str(cell.row)].value = "Market Value (Optimal)"
                    ws['AB'+str(cell.row)].value = "Shares"
                else:
                    print cell.row, cell.column,  cell.value, df.loc[cell.value]['longshort']
                    ws['N'+str(cell.row)].value = df.loc[cell.value]['longshort']
                    ws['U'+str(cell.row)].value = df.loc[cell.value]['forecastreturn']
                    ws['V'+str(cell.row)].value = df.loc[cell.value]['givenweight']
                    ws['Y'+str(cell.row)].value = df.loc[cell.value]['topconstraint']
                    ws['Z'+str(cell.row)].value = df.loc[cell.value]['bottomconstraint']
                    ws['W'+str(cell.row)].value = 1.0/float(len(df))
                    
        lastrow = cell.row
        ws['AC'+str(lastrow + 1)].value = 'Historical Return'
        ws['AC'+str(lastrow + 2)].value = 'Forecast Return'
        ws['AC'+str(lastrow + 3)].value = 'Stddev'
        ws['AC'+str(lastrow + 4)].value = 'Return/Risk (Historical)' 
        ws['AC'+str(lastrow + 5)].value = 'Return/Risk (Forecast)' 
        ws['AC'+str(lastrow + 6)].value = 'Annualized Stdev'
        
        ws_from = wb["permutations"]
        i0 = 0
        for col in ws_from.columns:
            i0 += 1
            #print 'col',col.column
            for idx, cell in enumerate(col, 1):
                ws.cell(row=i0, column=idx+30).value = cell.value #1-indexing
        #PortfolioReturn = MMULT(TRANSPOSE(totalreturn),weights)
        
        #Stdev =SQRT(MMULT(MMULT(TRANSPOSE(weights),covariancematrix,weights))
        cov_lastrow = len(wb["covariance"]['A'])       
        cov_lastcol = len(wb["covariance"][1])
        import xlsxwriter 
        cov_lastcolletter = xlsxwriter.utility.xl_col_to_name(cov_lastcol-1)
        #cell = xlsxwriter.utility.xl_rowcol_to_cell(1, 2)  # C2
        #print cell
        #stop

##        #Given
##        ws['V'+str(lastrow+1)].value = '=MMULT(TRANSPOSE(F2:F'+str(lastrow)+'),V2:V'+str(lastrow)+')' #return
##        ws['V'+str(lastrow+2)].value = '=MMULT(TRANSPOSE(U2:U'+str(lastrow)+'),V2:V'+str(lastrow)+')' #return
##        ws['V'+str(lastrow+3)].value = '=SQRT(MMULT(MMULT(TRANSPOSE('+'V2:V'+str(lastrow)+'),'+'covariance!B2:'+cov_lastcolletter+str(cov_lastrow)+'),'+'V2:V'+str(lastrow)+'))' #stdev
##        ws['V'+str(lastrow+4)].value = '=V'+str(lastrow+1)+'/V'+str(lastrow+3) #r/r historical
##        ws['V'+str(lastrow+5)].value = '=V'+str(lastrow+2)+'/V'+str(lastrow+3) #r/r forecast
##        ws['V'+str(lastrow+6)].value = '=V'+str(lastrow+3)+'*SQRT(250)'
        
        
        
        #from openpyxl.formula import Tokenizer
        #tok = Tokenizer("=MMULT(TRANSPOSE(F2:F"+str(lastrow)+"),"+weightscol+"2:"+weightscol+str(lastrow)+")")
        
        #tok.parse()

        #Given
        weightscol = 'V'
        ws[weightscol+str(lastrow+1)].value = '=MMULT(TRANSPOSE(F2:F'+str(lastrow)+'),'+weightscol+'2:'+weightscol+str(lastrow)+')' #return
        ws[weightscol+str(lastrow+2)].value = '=MMULT(TRANSPOSE(U2:U'+str(lastrow)+'),'+weightscol+'2:'+weightscol+str(lastrow)+')' #return
        ws[weightscol+str(lastrow+3)].value = '=SQRT(MMULT(MMULT(TRANSPOSE('+weightscol+'2:'+weightscol+str(lastrow)+'),'+'covariance!B2:'+cov_lastcolletter+str(cov_lastrow)+'),'+weightscol+'2'+':'+weightscol+str(lastrow)+'))' #stdev
        ws[weightscol+str(lastrow+4)].value = '='+weightscol+str(lastrow+1)+'/'+weightscol+str(lastrow+3) #r/r historical
        ws[weightscol+str(lastrow+5)].value = '='+weightscol+str(lastrow+2)+'/'+weightscol+str(lastrow+3) #r/r forecast
        ws[weightscol+str(lastrow+6)].value = '='+weightscol+str(lastrow+3)+'*SQRT(250)'
        
        #Equal
        weightscol = 'W'
        ws[weightscol+str(lastrow+1)].value = '=MMULT(TRANSPOSE(F2:F'+str(lastrow)+'),'+weightscol+'2:'+weightscol+str(lastrow)+')' #return
        ws[weightscol+str(lastrow+2)].value = '=MMULT(TRANSPOSE(U2:U'+str(lastrow)+'),'+weightscol+'2:'+weightscol+str(lastrow)+')' #return
        ws[weightscol+str(lastrow+3)].value = '=SQRT(MMULT(MMULT(TRANSPOSE('+weightscol+'2:'+weightscol+str(lastrow)+'),'+'covariance!B2:'+cov_lastcolletter+str(cov_lastrow)+'),'+weightscol+'2'+':'+weightscol+str(lastrow)+'))' #stdev
        ws[weightscol+str(lastrow+4)].value = '='+weightscol+str(lastrow+1)+'/'+weightscol+str(lastrow+3) #r/r historical
        ws[weightscol+str(lastrow+5)].value = '='+weightscol+str(lastrow+2)+'/'+weightscol+str(lastrow+3) #r/r forecast
        ws[weightscol+str(lastrow+6)].value = '='+weightscol+str(lastrow+3)+'*SQRT(250)'

        #Optimizable
        weightscol = 'X'
        ws[weightscol+str(lastrow+1)].value = '=MMULT(TRANSPOSE(F2:F'+str(lastrow)+'),'+weightscol+'2:'+weightscol+str(lastrow)+')' #return
        ws[weightscol+str(lastrow+2)].value = '=MMULT(TRANSPOSE(U2:U'+str(lastrow)+'),'+weightscol+'2:'+weightscol+str(lastrow)+')' #return
        ws[weightscol+str(lastrow+3)].value = '=SQRT(MMULT(MMULT(TRANSPOSE('+weightscol+'2:'+weightscol+str(lastrow)+'),'+'covariance!B2:'+cov_lastcolletter+str(cov_lastrow)+'),'+weightscol+'2'+':'+weightscol+str(lastrow)+'))' #stdev
        ws[weightscol+str(lastrow+4)].value = '='+weightscol+str(lastrow+1)+'/'+weightscol+str(lastrow+3) #r/r historical
        ws[weightscol+str(lastrow+5)].value = '='+weightscol+str(lastrow+2)+'/'+weightscol+str(lastrow+3) #r/r forecast
        ws[weightscol+str(lastrow+6)].value = '='+weightscol+str(lastrow+3)+'*SQRT(250)'


        weightscol = 'AE'
        i0 = 0
        while 1 == 1:
            #Might want to give this a shot: http://xlsxwriter.readthedocs.io/example_array_formula.html
            i0 += 1
            weightscol = ws[weightscol+'1'].offset(column=1).column
            print 'value',ws[weightscol+'1'].value
            if ws[weightscol+'1'].value == None:
                break
            ws[weightscol+str(lastrow+1)].value = '=MMULT(TRANSPOSE(F2:F'+str(lastrow)+'),'+weightscol+'2:'+weightscol+str(lastrow)+')' #return
            ws[weightscol+str(lastrow+2)].value = '=MMULT(TRANSPOSE(U2:U'+str(lastrow)+'),'+weightscol+'2:'+weightscol+str(lastrow)+')' #return
            ws[weightscol+str(lastrow+3)].value = '=SQRT(MMULT(MMULT(TRANSPOSE('+weightscol+'2:'+weightscol+str(lastrow)+'),'+'covariance!B2:'+cov_lastcolletter+str(cov_lastrow)+'),'+weightscol+'2'+':'+weightscol+str(lastrow)+'))' #stdev
            ws[weightscol+str(lastrow+4)].value = '='+weightscol+str(lastrow+1)+'/'+weightscol+str(lastrow+3) #r/r historical
            ws[weightscol+str(lastrow+5)].value = '='+weightscol+str(lastrow+2)+'/'+weightscol+str(lastrow+3) #r/r forecast
            ws[weightscol+str(lastrow+6)].value = '='+weightscol+str(lastrow+3)+'*SQRT(250)'
        print 'loop broken'
        
        wb.save(compiledfilepath)
        wb.close()
        print 'you can find your improved file here (PathnameToCompiledWorkbook): '
        print compiledfilepath

if __name__=='__main__':

    #df_symbols_and_signs = pd.read_csv('C:\\Batches\\GitStuff\\$work\\glse.csv')
    #df_symbols_and_signs = df_symbols_and_signs.set_index('ticker',drop=True)
    #print df_symbols_and_signs
    #stop
    lst0 = []
    lst0.append({'ticker':'MRK','longshort':'L','givenweight':0.2, 'forecastreturn':0.05,'topconstraint':0.05,'bottomconstraint':0.01})
    lst0.append({'ticker':'THO','longshort':'L','givenweight':0.2, 'forecastreturn':0.05,'topconstraint':0.05,'bottomconstraint':0.01})
    lst0.append({'ticker':'ALGN','longshort':'S','givenweight':0.2, 'forecastreturn':-0.05,'topconstraint':-0.005,'bottomconstraint':-0.02})
    lst0.append({'ticker':'CELG','longshort':'S','givenweight':0.2, 'forecastreturn':-0.05,'topconstraint':-0.005,'bottomconstraint':-0.02})
    lst0.append({'ticker':'MSFT','longshort':'L','givenweight':0.1, 'forecastreturn':0.05,'topconstraint':0.05,'bottomconstraint':0.01})
    lst0.append({'ticker':'FB','longshort':'L','givenweight':0.1, 'forecastreturn':0.05,'topconstraint':0.05,'bottomconstraint':0.01})
    df_symbols_and_signs = pd.DataFrame(lst0)
    df_symbols_and_signs = df_symbols_and_signs.set_index('ticker',drop=True)
    print df_symbols_and_signs
    #stop
    
    o = compileclass(
                symbols_and_signs_dataframe = df_symbols_and_signs
                ,  startdate = '2017-02-28'
                ,  enddate = '2017-10-24'
                ,  permutations = 500
                ,  annualized_or_cumulative = 'cumulative'
              )
  
    o.improveworkbook(symbols_and_signs_dataframe = df_symbols_and_signs,compiledfilepath=o.PathnameToCompiledWorkbook)
    
    #testpathname = 'C:\\Batches\\GitStuff\\modern-portfolio-theory\\cache\\20171023080411 compiled test.xlsx'
    #o.improveworkbook(symbols_and_signs_dataframe = df_symbols_and_signs,compiledfilepath=testpathname)
