import pandas as pd
class improve:
    

    def set_PathnameToCompiledWorkbook(self,PathnameToCompiledWorkbook):
        self._PathnameToCompiledWorkbook = PathnameToCompiledWorkbook
    def get_PathnameToCompiledWorkbook(self):
        return self._PathnameToCompiledWorkbook
    PathnameToCompiledWorkbook = property(get_PathnameToCompiledWorkbook, set_PathnameToCompiledWorkbook)

    
    def __init__(self,symbols_and_signs_dataframe,compiledfilepath):
        self.PathnameToCompiledWorkbook = compiledfilepath
        
        import config
        import mytools
        import openpyxl as pyxl
        import itertools
        wb = pyxl.load_workbook(compiledfilepath)

        #symbols = [i[0] for i in symbols_and_signs_list]
        #longshortlist = [i[1] for i in symbols_and_signs_list]
        df = symbols_and_signs_dataframe
        symbols = list(df.index)
        #print symbols
        #stop
        
        #for sh in wb.sheetnames:
        #    print sh
        #ws = wb["aggregatedpricechangereturns"]
        #ws = wb.copy_worksheet(ws)
        #ws.title = 'simulation'
        ws = wb["simulation"] #remove this later
        
        for row in ws.iter_rows('B{}:B{}'.format(ws.min_row,ws.max_row)):
            for cell in row:
                if cell.row == 1:
                    ws['N'+str(cell.row)].value = "Long/Short"
                    ws['U'+str(cell.row)].value = "Forecast Return"
                    ws['V'+str(cell.row)].value = "Given Weights"
                    ws['W'+str(cell.row)].value = "Equal Weights"
                    ws['X'+str(cell.row)].value = "Optimal Weights"
                    ws['Y'+str(cell.row)].value = "Top Constraint"
                    ws['Z'+str(cell.row)].value = "Bottom Constraint"
                    ws['AA'+str(cell.row)].value = "Market Value (Optimal)"
                    ws['AB'+str(cell.row)].value = "Shares"
                else:
                    print cell.row, cell.column,  cell.value, df.loc[cell.value]['longshort']
                    ws['N'+str(cell.row)].value = df.loc[cell.value]['longshort']
                    ws['U'+str(cell.row)].value = df.loc[cell.value]['forecastreturn']
                    ws['V'+str(cell.row)].value = df.loc[cell.value]['givenweight']
                    ws['Y'+str(cell.row)].value = df.loc[cell.value]['topconstraint']
                    ws['Z'+str(cell.row)].value = df.loc[cell.value]['bottomconstraint']
                    ws['W'+str(cell.row)].value = 1.0/float(len(df))
                    
                    
        #stop
        
        # Data can be assigned directly to cells
        #ws['N1'] = 42

        # Rows can also be appended
        #ws.append([1, 2, 3])

        # Python types will automatically be converted
        import datetime
        #ws['N2'] = datetime.datetime.now()

        # Save the file
        wb.save(compiledfilepath)
        wb.close()
        print 'you can find your compileclass file here (PathnameToCompiledWorkbook): ',compiledfilepath

if __name__=='__main__':
    lst0 = []
    lst0.append({'ticker':'MRK','longshort':'L','givenweight':0.2, 'forecastreturn':0.05,'topconstraint':0.05,'bottomconstraint':0.01})
    lst0.append({'ticker':'THO','longshort':'L','givenweight':0.2, 'forecastreturn':0.05,'topconstraint':0.05,'bottomconstraint':0.01})
    lst0.append({'ticker':'ALGN','longshort':'S','givenweight':0.2, 'forecastreturn':0.05,'topconstraint':-0.005,'bottomconstraint':-0.02})
    lst0.append({'ticker':'CELG','longshort':'S','givenweight':0.2, 'forecastreturn':0.05,'topconstraint':-0.005,'bottomconstraint':-0.02})
    lst0.append({'ticker':'MSFT','longshort':'L','givenweight':0.1, 'forecastreturn':0.05,'topconstraint':0.05,'bottomconstraint':0.01})
    lst0.append({'ticker':'FB','longshort':'L','givenweight':0.1, 'forecastreturn':0.05,'topconstraint':0.05,'bottomconstraint':0.01})
    df_symbols_and_signs = pd.DataFrame(lst0)
    df_symbols_and_signs = df_symbols_and_signs.set_index('ticker',drop=True)
    print df_symbols_and_signs
    #stop

    o = improve( symbols_and_signs_dataframe = df_symbols_and_signs
            , compiledfilepath='C:\\Batches\\GitStuff\\modern-portfolio-theory\\cache\\20171019152219 compiled.xlsx'
            
               )
